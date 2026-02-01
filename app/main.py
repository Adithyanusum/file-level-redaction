from fastapi import FastAPI, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, RedirectResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
import io, json
from . import redact

app = FastAPI(title="File Redaction Hackathon")
app.mount("/static", StaticFiles(directory="static"), name="static")

# allow requests from local dev servers (e.g. Live Server on :5500)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # permissive for local testing
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Redacted", "X-Regions-Count", "X-Phrases-Count", "X-First-Region", "Content-Disposition"],
)


@app.on_event("startup")
def on_startup():
    print("[app] startup complete")


@app.on_event("shutdown")
def on_shutdown():
    print("[app] shutdown")


@app.get('/health')
def health():
    return JSONResponse({'status': 'ok'})


@app.post("/redact/image")
async def redact_image(file: UploadFile = File(...), regions: str = Form(None), phrases: str = Form(None), mode: str = Form("blackout")):
    data = await file.read()
    regions_list = json.loads(regions) if regions else []
    phrases_list = json.loads(phrases) if phrases else []
    print(f"[redact_image] filename={file.filename} mode={mode} regions={regions_list} phrases={phrases_list}")
    # If phrases provided, attempt to locate them via OCR (server-side) and add to regions
    try:
        if phrases_list:
            # detect words/boxes in image using pytesseract (if available)
            try:
                matches = redact.detect_image_bytes(data)
                if isinstance(matches, dict):
                    if matches.get('error'):
                        matches = []
                    else:
                        matches = matches.get('matches', [])
            except Exception:
                matches = []
            for ph in phrases_list:
                if not ph: continue
                for m in (matches or []):
                    try:
                        txt = (m.get('text') or '')
                        if txt and ph.lower() in txt.lower():
                            rect = m.get('rect')
                            if rect and len(rect) == 4:
                                regions_list.append(rect)
                    except Exception:
                        continue
    except Exception:
        pass
    out_bytes = redact.redact_image_bytes(data, regions_list, mode)
    headers = {"Content-Disposition": f'attachment; filename="redacted-{file.filename}"'}
    return StreamingResponse(io.BytesIO(out_bytes), media_type=file.content_type, headers=headers)


@app.post("/redact/pdf")
async def redact_pdf(file: UploadFile = File(...), regions: str = Form(None), phrases: str = Form(None)):
    try:
        data = await file.read()
        # log raw incoming form values for debug
        print(f"[redact_pdf] raw regions_str={regions}")
        print(f"[redact_pdf] raw phrases_str={phrases}")
        regions_obj = json.loads(regions) if regions else []
        phrases_list = json.loads(phrases) if phrases else []
        # If phrases provided, search PDF pages for those phrases and add to regions
        if phrases_list:
            try:
                doc = redact.fitz.open(stream=data, filetype='pdf')
                for pno in range(len(doc)):
                    page = doc.load_page(pno)
                    for ph in phrases_list:
                            try:
                                found_any = False
                                # try direct search first
                                try:
                                    areas = page.search_for(ph)
                                    for r in areas:
                                        regions_obj.append({"page": pno, "rect": [r.x0, r.y0, r.x1, r.y1]})
                                        found_any = True
                                except Exception:
                                    pass
                                if not found_any:
                                    # fallback: match using word boxes (case-insensitive)
                                    try:
                                        words = page.get_text("words")  # list of tuples: x0,y0,x1,y1,word
                                        import re as _re
                                        # normalize phrase
                                        phrase_norm = _re.sub(r"\s+", " ", _re.sub(r"[^\w\s]", " ", ph)).strip().lower()
                                        if words and phrase_norm:
                                            # build list of normalized words
                                            norm_words = [_re.sub(r"[^\w]", "", w[4]).lower() for w in words]
                                            pw = [pw for pw in phrase_norm.split() if pw]
                                            plen = len(pw)
                                            if plen == 0:
                                                pass
                                            else:
                                                for i in range(0, max(0, len(norm_words) - 0)):
                                                    # build a window string from i to i+max_window
                                                    for end in range(i + 1, min(len(norm_words), i + plen + 6) + 1):
                                                        seq = norm_words[i:end]
                                                        joined = " ".join(seq)
                                                        if pw == seq or phrase_norm in joined or joined in phrase_norm:
                                                            # compute bounding box across words i..end-1
                                                            try:
                                                                x0 = min(words[k][0] for k in range(i, end))
                                                                y0 = min(words[k][1] for k in range(i, end))
                                                                x1 = max(words[k][2] for k in range(i, end))
                                                                y1 = max(words[k][3] for k in range(i, end))
                                                                regions_obj.append({"page": pno, "rect": [x0, y0, x1, y1]})
                                                                found_any = True
                                                                break
                                                            except Exception:
                                                                continue
                                                    if found_any:
                                                        break
                                    except Exception:
                                        pass
                            except Exception:
                                continue
                doc.close()
            except Exception:
                pass
        print(f"[redact_pdf] filename={file.filename} parsed_regions={regions_obj} phrases={phrases_list}")
        # Normalize incoming regions format:
        # - If client sent a list of simple [x,y,w,h] arrays, convert to {page:0, rect:[x0,y0,x1,y1]}
        try:
            norm_regions = []
            if isinstance(regions_obj, list) and regions_obj:
                # detect simple array format (canvas pixel boxes like [x,y,w,h])
                if isinstance(regions_obj[0], list) and (len(regions_obj[0]) >= 4) and not isinstance(regions_obj[0][0], dict):
                    try:
                        # convert canvas pixel boxes into per-page PDF rects assuming preview used zoom=2.0
                        zoom = 2.0
                        import fitz as _fitz
                        _doc = _fitz.open(stream=data, filetype='pdf')
                        page_pix_heights = []
                        for pno in range(len(_doc)):
                            pg = _doc.load_page(pno)
                            pix = pg.get_pixmap(matrix=_fitz.Matrix(zoom, zoom))
                            page_pix_heights.append(pix.height)
                        # cumulative heights to locate page by y coordinate
                        cum = [0]
                        for hgt in page_pix_heights:
                            cum.append(cum[-1] + hgt)
                        total_pages = len(page_pix_heights)
                        for r in regions_obj:
                            try:
                                x = float(r[0]); y = float(r[1]); w = float(r[2]); h = float(r[3])
                                # find page index by y coordinate
                                pidx = 0
                                for i in range(total_pages):
                                    if y >= cum[i] and y < cum[i+1]:
                                        pidx = i; break
                                y_in_page = y - cum[pidx]
                                x0 = x / zoom; y0 = y_in_page / zoom; x1 = (x + w) / zoom; y1 = (y_in_page + h) / zoom
                                norm_regions.append({"page": pidx, "rect": [x0, y0, x1, y1]})
                            except Exception:
                                continue
                        try: _doc.close()
                        except Exception: pass
                    except Exception:
                        # fallback to naive conversion if anything fails
                        for r in regions_obj:
                            try:
                                x = float(r[0]); y = float(r[1]); w = float(r[2]); h = float(r[3])
                                norm_regions.append({"page": 0, "rect": [x, y, x + w, y + h]})
                            except Exception:
                                continue
                else:
                    # assume already list of dicts or proper format
                    for item in regions_obj:
                        if isinstance(item, dict) and item.get('rect'):
                            norm_regions.append(item)
            regions_obj = norm_regions
            # If still no regions, fallback to server-side detection
            if not regions_obj:
                detected = redact.detect_pdf_bytes(data)
                if detected:
                    for pg in detected:
                        pno = pg.get('page', 0)
                        for m in (pg.get('matches') or []):
                            rect = m.get('rect')
                            if rect:
                                regions_obj.append({"page": pno, "rect": rect})
                    print(f"[redact_pdf] auto-detected regions count={len(regions_obj)}")
        except Exception as e:
            print(f"[redact_pdf] normalize/detect fallback error: {e}")
        out_bytes = redact.redact_pdf_bytes(data, regions_obj)
        modified = (out_bytes != data)
        # expose debug headers: count of regions and first region JSON (if small)
        rcount = len(regions_obj) if regions_obj else 0
        pcount = len(phrases_list) if phrases_list else 0
        first_region = json.dumps(regions_obj[0]) if rcount>0 else ""
        headers = {
            "Content-Disposition": f'attachment; filename="redacted-{file.filename}"',
            "X-Redacted": ("true" if modified else "false"),
            "X-Regions-Count": str(rcount),
            "X-Phrases-Count": str(pcount),
            "X-First-Region": first_region
        }
        print(f"[redact_pdf] filename={file.filename} modified={modified} regions_count={rcount} phrases_count={pcount}")
        # Return as octet-stream to encourage download in browsers
        return StreamingResponse(io.BytesIO(out_bytes), media_type="application/octet-stream", headers=headers)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[redact_pdf] ERROR: {e}\n{tb}")
        return JSONResponse({"error": str(e), "trace": tb}, status_code=500)


@app.post("/redact/docx")
async def redact_docx(file: UploadFile = File(...), phrases: str = Form(None), media_to_blur: str = Form(None)):
    data = await file.read()
    phrases_list = json.loads(phrases) if phrases else []
    media_list = json.loads(media_to_blur) if media_to_blur else None
    out_bytes = redact.redact_docx_bytes(data, phrases_list, media_to_blur=media_list)
    headers = {"Content-Disposition": f'attachment; filename="redacted-{file.filename}"'}
    return StreamingResponse(io.BytesIO(out_bytes), media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", headers=headers)


@app.post("/redact/xlsx")
async def redact_xlsx(file: UploadFile = File(...), cells: str = Form(None), columns: str = Form(None), rows: str = Form(None), phrases: str = Form(None), media_to_blur: str = Form(None)):
    data = await file.read()
    cells_list = json.loads(cells) if cells else []
    columns_list = json.loads(columns) if columns else []
    rows_list = json.loads(rows) if rows else []
    phrases_list = json.loads(phrases) if phrases else []
    media_list = json.loads(media_to_blur) if media_to_blur else None
    out_bytes = redact.redact_xlsx_bytes(data, cells_list, columns_list, rows_list, phrases=phrases_list, media_to_blur=media_list)
    headers = {"Content-Disposition": f'attachment; filename="redacted-{file.filename}"'}
    return StreamingResponse(io.BytesIO(out_bytes), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@app.get("/")
def home():
    return RedirectResponse(url='/static/index.html')


@app.post('/preview/pdf')
async def preview_pdf(file: UploadFile = File(...)):
    data = await file.read()
    try:
        out = redact.preview_pdf_first_page(data)
        return StreamingResponse(io.BytesIO(out), media_type='image/png')
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)


@app.post('/preview/docx')
async def preview_docx(file: UploadFile = File(...), format: str = Form(None)):
    data = await file.read()
    try:
        if format == 'html':
            html_out = redact.preview_docx_html(data)
            return HTMLResponse(content=html_out)
        out = redact.preview_docx_bytes(data)
        return StreamingResponse(io.BytesIO(out), media_type='image/png')
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)


@app.post('/preview/xlsx')
async def preview_xlsx(file: UploadFile = File(...), format: str = Form(None), max_rows: str = Form(None), max_cols: str = Form(None)):
    data = await file.read()
    try:
        if format == 'html':
            # parse optional max limits from form; if not provided, show all rows
            mr = int(max_rows) if (max_rows and str(max_rows).isdigit()) else None
            mc = int(max_cols) if (max_cols and str(max_cols).isdigit()) else None
            html_out = redact.preview_xlsx_html(data, max_rows=mr, max_cols=mc)
            return HTMLResponse(content=html_out)
        out = redact.preview_xlsx_bytes(data)
        return StreamingResponse(io.BytesIO(out), media_type='image/png')
    except Exception as e:
        return JSONResponse({'error': str(e)}, status_code=500)


@app.post('/detect')
async def detect(file: UploadFile = File(...)):
    try:
        data = await file.read()
        name = file.filename.lower()
        if name.endswith('.pdf'):
            res = redact.detect_pdf_bytes(data)
            return JSONResponse(res)
        if name.endswith('.docx'):
            res = redact.detect_docx_bytes(data)
            return JSONResponse(res)
        if name.endswith('.xlsx'):
            res = redact.detect_xlsx_bytes(data)
            return JSONResponse(res)
        # image
        res = redact.detect_image_bytes(data)
        # detect_image_bytes may return {'matches': [...], 'full_text': '...'} or an error dict
        if isinstance(res, dict) and ('matches' in res or 'error' in res):
            return JSONResponse(res)
        return JSONResponse({'matches': res})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[detect] ERROR: {e}\n{tb}")
        return JSONResponse({'error': str(e), 'trace': tb}, status_code=500)


@app.post('/extract')
async def extract_text(file: UploadFile = File(...)):
    """Return extracted full text for a file (pdf, image, docx, xlsx) to populate the 'Redact more' panel."""
    try:
        data = await file.read()
        name = file.filename.lower()
        if name.endswith('.pdf'):
            # concatenate page text
            try:
                doc = redact.fitz.open(stream=data, filetype='pdf')
                parts = []
                for pno in range(len(doc)):
                    try:
                        page = doc.load_page(pno)
                        parts.append(page.get_text())
                    except Exception:
                        continue
                doc.close()
                return JSONResponse({'full_text': '\n'.join(parts)})
            except Exception:
                return JSONResponse({'full_text': ''})
        if name.endswith('.docx'):
            try:
                res = redact.detect_docx_bytes(data)
                # return textual matches and attempt to include all paragraphs
                buf = io.BytesIO(data)
                from docx import Document
                doc = Document(buf)
                paras = []
                for p in doc.paragraphs:
                    paras.append(p.text)
                return JSONResponse({'full_text': '\n'.join(paras)})
            except Exception:
                return JSONResponse({'full_text': ''})
        if name.endswith('.xlsx'):
            try:
                # return a simple CSV-like text of the first sheet
                from openpyxl import load_workbook
                buf = io.BytesIO(data)
                wb = load_workbook(filename=buf, data_only=True)
                ws = wb.active
                rows = []
                for r in ws.iter_rows(values_only=True):
                    rows.append('\t'.join(['' if c is None else str(c) for c in r]))
                return JSONResponse({'full_text': '\n'.join(rows)})
            except Exception:
                return JSONResponse({'full_text': ''})
        # image: try server-side pytesseract if available, else return empty so client can fallback
        text = ''
        try:
            if redact.pytesseract is not None:
                from PIL import Image
                img = Image.open(io.BytesIO(data))
                text = redact.pytesseract.image_to_string(img)
        except Exception:
            text = ''
        return JSONResponse({'full_text': text})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[extract] ERROR: {e}\n{tb}")
        return JSONResponse({'full_text': ''})


@app.post('/redact/auto')
async def redact_auto(file: UploadFile = File(...), mode: str = Form('blackout')):
    """Detect and redact all sensitive data found in the uploaded file automatically."""
    try:
        data = await file.read()
        name = file.filename.lower()
        # PDF: detect text matches and convert to regions, then redact
        if name.endswith('.pdf'):
            detected = redact.detect_pdf_bytes(data)
            regions = []
            if detected:
                for pg in detected:
                    pno = pg.get('page', 0)
                    for m in (pg.get('matches') or []):
                        rect = m.get('rect')
                        if rect:
                            regions.append({"page": pno, "rect": rect})
            out = redact.redact_pdf_bytes(data, regions)
            headers = {"Content-Disposition": f'attachment; filename="redacted-{file.filename}"'}
            return StreamingResponse(io.BytesIO(out), media_type='application/pdf', headers=headers)
        # DOCX: extract text, scan for sensitive phrases, and redact via redact_docx_bytes
        if name.endswith('.docx'):
            # Detect text and embedded image matches, then redact text phrases and selectively blur matching images
            detected = redact.detect_docx_bytes(data)
            phrases = []
            for m in (detected.get('text_matches') or []):
                if isinstance(m, dict):
                    phrases.append(m.get('match'))
                else:
                    phrases.append(m)
            # collect embedded media basenames that had matches
            media_to_blur = []
            for im in (detected.get('image_matches') or []):
                try:
                    nm = im.get('image')
                    if nm:
                        media_to_blur.append(nm)
                except Exception:
                    continue
            # dedupe phrases
            phrases = [p for p in dict.fromkeys([p for p in phrases if p])]
            out = redact.redact_docx_bytes(data, phrases, media_to_blur=media_to_blur)
            headers = {"Content-Disposition": f'attachment; filename="redacted-{file.filename}"'}
            return StreamingResponse(io.BytesIO(out), media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document", headers=headers)
        # XLSX: scan sheet text for sensitive phrases and redact via redact_xlsx_bytes
        if name.endswith('.xlsx'):
            detected = redact.detect_xlsx_bytes(data)
            phrases = []
            for item in (detected.get('text_matches') or []):
                if isinstance(item, dict):
                    phrases.append(item.get('match'))
                else:
                    phrases.append(item)
            media_to_blur = []
            for im in (detected.get('image_matches') or []):
                try:
                    nm = im.get('image')
                    if nm:
                        media_to_blur.append(nm)
                except Exception:
                    continue
            # dedupe
            phrases = [p for p in dict.fromkeys([p for p in phrases if p])]
            out = redact.redact_xlsx_bytes(data, cells=[], columns=[], rows=None, phrases=phrases, media_to_blur=media_to_blur)
            headers = {"Content-Disposition": f'attachment; filename="redacted-{file.filename}"'}
            return StreamingResponse(io.BytesIO(out), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
        # Image: try server OCR to detect matches and redact by drawing boxes
        if name.endswith('.png') or name.endswith('.jpg') or name.endswith('.jpeg') or name.endswith('.tiff') or name.endswith('.bmp'):
            matches = redact.detect_image_bytes(data)
            rects = []
            if isinstance(matches, dict):
                matches = matches.get('matches', [])
            for m in (matches or []):
                r = m.get('rect') if isinstance(m, dict) else None
                if r and len(r) == 4:
                    # ensure format [x,y,w,h]
                    # server detect_image_bytes returns [x,y,w,h]
                    rects.append(r)
                elif r and len(r) == 4:
                    # fallback if rect was [x0,y0,x1,y1]
                    rects.append([r[0], r[1], r[2]-r[0], r[3]-r[1]])
            out = redact.redact_image_bytes(data, rects, mode)
            headers = {"Content-Disposition": f'attachment; filename="redacted-{file.filename}"'}
            return StreamingResponse(io.BytesIO(out), media_type='image/png', headers=headers)
        # fallback: return original
        return JSONResponse({'error': 'unsupported file type for auto redact'}, status_code=400)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[redact/auto] ERROR: {e}\n{tb}")
        return JSONResponse({'error': str(e), 'trace': tb}, status_code=500)
